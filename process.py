"""
process.py — CE East master pipeline.

Reads:
  incoming/*.xlsx   — CE East load book (source of truth: status, $, release flag)
  incoming/*.pdf    — NEW Flexent reserve detail reports (cross-check + release dates)
  docs/*.pdf        — PREVIOUSLY archived Flexent PDFs, re-read every run so PDF-only
                      release flags never drop off when a new weekly report is added.

Writes:
  public/data/loads.csv          — every CE East load, with `released` derived from
                                   the Excel "RELEASED RESERVES" column AND any Pmt
                                   row found in the PDFs. PDF dates fill release_date.
  public/data/reserve_status.csv — invoice-level release events parsed out of PDFs
                                   (kept for cross-check / audit; the dashboard reads
                                   loads.csv only).

Side effects:
  New PDFs in incoming/ are moved to docs/ after parsing. PDFs already in docs/ stay put.
  Excel files stay in incoming/ — they're live, the user keeps updating them.

Usage:
  pip install -r requirements.txt
  python process.py

Then commit + push to deploy.
"""

from __future__ import annotations

import csv
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import pdfplumber

BASE = Path(__file__).resolve().parent
INCOMING = BASE / "incoming"
ARCHIVE = BASE / "docs"
OUT_LOADS = BASE / "public" / "data" / "loads.csv"
OUT_RESERVES = BASE / "public" / "data" / "reserve_status.csv"

LOAD_SHEET_HINTS = ("invoice list", "inv")  # any sheet whose name contains these is a load sheet

# Reserve rate Flexent holds on every funded invoice. Used to recompute reserves
# per-load — overrides the Excel RESERVES column, which has stale rates from older
# fee schedules and is missing on many rows.
RESERVE_RATE = 0.05

# Customer name normalizer: collapse common spelling/location variants.
# `(prefix, canonical)` — any name whose normalized form starts with `prefix` becomes `canonical`.
# Order matters: longer / more-specific prefixes first.
CUSTOMER_PREFIX_RULES = [
    ("rentex",                 "Rentex"),
    ("on services",            "On Services"),
    ("on-services",            "On Services"),
    ("gofo express",           "Gofo Express"),
    ("gofo inc",               "Gofo Inc"),
    ("gofo",                   "Gofo Inc"),
    ("insync",                 "Insync Productions"),
    ("insomniac",              "Insomniac"),
    ("sierra live productions","Sierra Live Productions"),
    ("iron clad",              "Ironclad Environmental"),
    ("ironclad",               "Ironclad Environmental"),
    ("freeman audio visual",   "Freeman Audio Visual"),
    ("4wall entertainment",    "4Wall Entertainment"),
    ("capacity express",       "Capacity Express"),
    ("capacity",               "Capacity Express"),
]


def normalize_customer(name) -> str:
    """Collapse customer name variants (Rentex - MA, RENTEX MASSACHUSETTS, etc.) to a canonical form."""
    if name is None:
        return ""
    s = str(name).strip()
    if not s:
        return ""
    # Lowercase + strip common punctuation + leading flags + suffixes for matching
    n = s.lower()
    # Strip leading flags like "(dnu)" (do not use), "(old)", "(new)" — these are notes, not part of the name
    n = re.sub(r"^\s*\(\s*(dnu|old|new|do not use)\s*\)\s*", "", n)
    n = re.sub(r"[,./]", " ", n)
    n = re.sub(r"\b(inc|llc|corp|company|co|hq|headquarters)\b\.?", "", n)
    n = re.sub(r"[-_]+", " ", n)
    n = re.sub(r"\s+", " ", n).strip()
    for prefix, canonical in CUSTOMER_PREFIX_RULES:
        if n.startswith(prefix):
            return canonical
    # No rule matched — return the original with light cleanup (title case, no extra spaces)
    return re.sub(r"\s+", " ", s.strip()).title()

# ── Excel helpers ──────────────────────────────────────────────────

def _norm(s: object) -> str:
    return str(s or "").strip().lower()


_HEADER_TARGETS = {"invoice", "invoice amount", "funded", "reserves", "invno", "amt", "released reserves", "reserves released"}


def _row_is_header(row_cells: list[str]) -> dict[str, int] | None:
    """If this row looks like a load-table header, return the mapping; else None."""
    if sum(1 for c in row_cells if c in _HEADER_TARGETS) < 3:
        return None
    return {c: i for i, c in enumerate(row_cells) if c}


def _to_iso(v) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return s


def _to_num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("$", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def _row_to_load(row, mapping, sheet):
    g = lambda *keys: next((row[mapping[k]] for k in keys if k in mapping and mapping[k] < len(row)), None)

    inv = g("invoice", "invno")
    if inv is None or inv == "":
        return None
    # Skip footer rows ("Totals") and embedded sub-headers ("InvNo", "Invoice")
    if isinstance(inv, str):
        try:
            float(inv.replace(",", ""))
        except ValueError:
            return None
    inv_str = str(int(inv)) if isinstance(inv, float) and inv.is_integer() else str(inv).strip()

    released_reserves = _to_num(g("released reserves", "reserves released"))
    invoice_amount = _to_num(g("invoice amount", "amt"))
    reserve_held = round(invoice_amount * RESERVE_RATE, 2) if invoice_amount else None
    is_released = bool(released_reserves and released_reserves > 0)
    return {
        "invoice_number": inv_str,
        "status": str(g("status") or "").strip(),
        "invoice_date": _to_iso(g("invoice date", "invdate")),
        "submission_date": _to_iso(g("submission date")),
        "po_load": str(g("po/load", "pono") or "").strip(),
        "customer": normalize_customer(g("customer", "debtor name")),
        "customer_raw": str(g("customer", "debtor name") or "").strip(),
        "age": _to_num(g("age")),
        "invoice_amount": invoice_amount,
        "funded": _to_num(g("funded")),
        "reserves": reserve_held,                 # 5% of invoice — what Flexent actually holds
        "reserves_excel": _to_num(g("reserves")), # raw Excel value (for audit)
        "released_reserves": released_reserves,
        "fees": _to_num(g("fees")),
        "carrier": str(g("carrier", "carrier name") or "").strip(),
        "carrier_invoice": str(g("carrier invoice") or "").strip(),
        "carrier_pay": _to_num(g("carrier invoice amount", "carrier pay")),
        "source_sheet": sheet,
        # release fields filled below from PDF cross-check.
        # Excel `RELEASED RESERVES` keying is treated as a Triumph-era signal only
        # (loads released under our prior factor, before Flexent). It's overridden
        # below whenever a Flexent PDF Pmt exists for the invoice.
        "released": "true" if is_released else "false",
        "release_date": "",
        # If released, the amount released equals the reserve held (5% of invoice)
        "release_amount": reserve_held if is_released else "",
        "release_source": "triumph" if is_released else "",
    }


def parse_excel(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    out = []
    for sheet_name in wb.sheetnames:
        if not any(h in sheet_name.lower() for h in LOAD_SHEET_HINTS):
            continue
        ws = wb[sheet_name]
        mapping: dict[str, int] | None = None
        for row in ws.iter_rows(values_only=True):
            cells = [_norm(c) for c in row]
            new_mapping = _row_is_header(cells)
            if new_mapping is not None:
                mapping = new_mapping
                continue
            if mapping is None:
                continue
            load = _row_to_load(row, mapping, sheet_name)
            if load:
                out.append(load)
    return out


# ── PDF helpers ──────────────────────────────────────────────────

# Match a Flexent reserve activity Pmt row. Layout:
#   <invoice>  <po/ref>  <check>  <debtor (multi-word)>  <date>  <days>  Pmt  <amount>  ...  <reserve_amt>  <balance>
PMT_RE = re.compile(
    r"^(?P<invoice>\S+)\s+\S+\s+\S+\s+.*?\s+(?P<date>\d{1,2}/\d{1,2}/\d{4})\s+\d+\s+Pmt\s+(?P<amount>[\d,]+\.\d{2})"
)


def parse_pdf(pdf_path: Path) -> list[dict]:
    rows = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.splitlines():
                m = PMT_RE.match(line)
                if not m:
                    continue
                inv = m.group("invoice")
                # Strip REV suffix; keep base invoice for join
                base_inv = re.sub(r"REV$", "", inv)
                rows.append({
                    "invoice_number": base_inv,
                    "released": "true",
                    "release_date": _to_iso(m.group("date")),
                    "release_amount": m.group("amount").replace(",", ""),
                    "source_pdf": pdf_path.name,
                })
    return rows


# ── Main ──────────────────────────────────────────────────

def main() -> int:
    if not INCOMING.exists():
        print(f"no incoming/ folder at {INCOMING}")
        return 1

    OUT_LOADS.parent.mkdir(parents=True, exist_ok=True)

    # 1) Excel -> loads (deduped by invoice_number, last sheet wins)
    xlsx_files = sorted(INCOMING.glob("*.xlsx"))
    if not xlsx_files:
        print("warning: no .xlsx in incoming/")
    by_inv: dict[str, dict] = {}
    parsed_total = 0
    for xlsx in xlsx_files:
        rows = parse_excel(xlsx)
        parsed_total += len(rows)
        for r in rows:
            by_inv[r["invoice_number"]] = r  # last write wins; monthly sheets supersede 2025 list
        print(f"  excel: {xlsx.name} -> {len(rows)} rows")
    loads = list(by_inv.values())
    if parsed_total != len(loads):
        print(f"  note:  deduped {parsed_total - len(loads)} duplicate invoice rows")

    # 2) PDFs -> release events (cross-check + dates)
    # Read NEW PDFs from incoming/ AND every archived PDF from docs/ so PDF-only
    # release flags persist across weeks — Flexent reports are cumulative-only for
    # the current period, so once a Pmt row drops off a newer report we'd lose it
    # if we didn't keep re-parsing the archive.
    ARCHIVE.mkdir(exist_ok=True)
    new_pdfs = sorted(INCOMING.glob("*.pdf"))
    archived_pdfs = sorted(ARCHIVE.glob("*.pdf"))
    pdf_releases: dict[str, dict] = {}

    def _ingest(pdf: Path, archive_after: bool) -> None:
        try:
            rows = parse_pdf(pdf)
        except Exception as e:
            print(f"  ! {pdf.name}: {e}")
            return
        new_count = 0
        for r in rows:
            if r["invoice_number"] not in pdf_releases:
                new_count += 1
            pdf_releases[r["invoice_number"]] = r
        label = "pdf-new" if archive_after else "pdf-arc"
        print(f"  {label}: {pdf.name} -> {len(rows)} Pmt rows ({new_count} unique)")
        if archive_after:
            shutil.move(str(pdf), ARCHIVE / pdf.name)

    for pdf in archived_pdfs:
        _ingest(pdf, archive_after=False)
    for pdf in new_pdfs:
        _ingest(pdf, archive_after=True)

    # 3) Cross-merge: any load with a Flexent PDF Pmt is released-by-flexent (PDF is
    # authoritative for the Flexent era — Ben's Excel keying is incidental). Loads
    # without a PDF Pmt fall back to the Triumph-era Excel flag set in step 1.
    flexent_count = 0
    triumph_count = 0
    for load in loads:
        inv = load["invoice_number"]
        rel = pdf_releases.get(inv)
        if rel:
            load["released"] = "true"
            load["release_amount"] = load["reserves"] or ""
            load["release_source"] = "flexent"
            load["release_date"] = rel["release_date"]
            flexent_count += 1
        elif load["release_source"] == "triumph":
            triumph_count += 1
    print(f"  note:  {flexent_count} released by Flexent (PDF), {triumph_count} released by Triumph (pre-Flexent, Excel only)")

    # 4) Write loads.csv
    fieldnames = [
        "invoice_number", "status", "invoice_date", "submission_date", "po_load",
        "customer", "customer_raw", "age", "invoice_amount", "funded", "reserves", "reserves_excel",
        "released_reserves", "fees", "carrier", "carrier_invoice", "carrier_pay", "source_sheet",
        "released", "release_date", "release_amount", "release_source",
    ]
    with OUT_LOADS.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in loads:
            w.writerow({k: ("" if v is None else v) for k, v in r.items()})
    print(f"wrote {OUT_LOADS.relative_to(BASE)} ({len(loads)} loads)")

    # 5) Write reserve_status.csv (PDF-only, audit/cross-check artifact)
    with OUT_RESERVES.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["invoice_number", "released", "release_date", "release_amount", "source_pdf"])
        w.writeheader()
        for inv in sorted(pdf_releases):
            w.writerow(pdf_releases[inv])
    print(f"wrote {OUT_RESERVES.relative_to(BASE)} ({len(pdf_releases)} PDF release events)")

    # 6) Summary
    released = sum(1 for r in loads if r["released"] == "true")
    unreleased = len(loads) - released
    print(f"\nsummary: {len(loads)} loads · {released} released · {unreleased} unreleased")
    return 0


if __name__ == "__main__":
    sys.exit(main())
